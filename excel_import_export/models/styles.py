# Copyright 2019 Ecosoft Co., Ltd (http://ecosoft.co.th/)
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl.html)

from odoo import models, api
import logging

_logger = logging.getLogger(__name__)

try:
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
except ImportError:
    _logger.debug(
        'Cannot import "openpyxl". Please make sure it is installed.')


class XLSXStyles(models.AbstractModel):
    _name = 'xlsx.styles'
    _description = 'Available styles for excel'

    @api.model
    def get_openpyxl_styles(self):
        """ List all syles that can be used with styleing directive #{...} """
        return {
            'font': {
                'bold': Font(name="Arial", size=10, bold=True),
                'bold_red': Font(name="Arial", size=10,
                                 color="FF0000", bold=True),
            },
            'fill': {
                'red': PatternFill("solid", fgColor="FF0000"),
                'grey': PatternFill("solid", fgColor="DDDDDD"),
                'yellow': PatternFill("solid", fgColor="FFFCB7"),
                'blue': PatternFill("solid", fgColor="9BF3FF"),
                'green': PatternFill("solid", fgColor="B0FF99"),
            },
            'align': {
                'left': Alignment(horizontal='left'),
                'center': Alignment(horizontal='center'),
                'right': Alignment(horizontal='right'),
                'full_center': Alignment(horizontal='center', vertical='center')
            },
            'vert-align': {
                'top': Alignment(vertical='top'),
                'justify': Alignment(vertical='justify'),
                'center': Alignment(vertical='center'),
                'bottom': Alignment(vertical='bottom'),
            },
            'border': {
                'thin_border': Border(left=Side(style='thin'),
                                      right=Side(style='thin'),
                                      top=Side(style='thin'),
                                      bottom=Side(style='thin'))
            },
            'style': {
                'number': '#,##0.00',
                'date': 'dd/mm/yyyy',
                'datestamp': 'yyyy-mm-dd',
                'text': '@',
                'percent': '0.00%',
            },
        }
